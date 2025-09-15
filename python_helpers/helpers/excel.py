#!/usr/bin/env python3
"""
excel.py - Excel backend for Excel::Writer::XLSX replacement

Provides Excel functionality using openpyxl or xlsxwriter.
Focused implementation based on actual usage analysis from enterprise Perl scripts.
"""

import uuid
from typing import Dict, Any, Optional

# Global workbook storage
WORKBOOKS = {}
WORKSHEETS = {}
FORMATS = {}

def create_workbook(filename: str) -> Dict[str, Any]:
    """
    Create a new Excel workbook
    
    Args:
        filename: Path to save the Excel file
    
    Returns:
        Dictionary with workbook creation result and ID
    """
    try:
        # Try openpyxl first (most compatible)
        try:
            from openpyxl import Workbook
            return _create_workbook_openpyxl(filename)
        except ImportError:
            # Fall back to xlsxwriter
            try:
                import xlsxwriter
                return _create_workbook_xlsxwriter(filename)
            except ImportError:
                # Fall back to basic CSV-like implementation
                return _create_workbook_basic(filename)
                
    except Exception as e:
        return {
            'success': False,
            'error': f'Excel workbook creation failed: {str(e)}',
        }

def _create_workbook_openpyxl(filename: str) -> Dict[str, Any]:
    """Create workbook using openpyxl"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border
    
    workbook = Workbook()
    # Remove default sheet - we'll add our own
    if workbook.worksheets:
        workbook.remove(workbook.active)
    
    workbook_id = str(uuid.uuid4())
    WORKBOOKS[workbook_id] = {
        'workbook': workbook,
        'filename': filename,
        'library': 'openpyxl',
        'worksheets': {},
        'formats': {}
    }
    
    return {
        'success': True,
        'result': {
            'workbook_id': workbook_id,
            'library': 'openpyxl',
            'filename': filename
        }
    }

def _create_workbook_xlsxwriter(filename: str) -> Dict[str, Any]:
    """Create workbook using xlsxwriter"""
    import xlsxwriter
    
    workbook = xlsxwriter.Workbook(filename)
    
    workbook_id = str(uuid.uuid4())
    WORKBOOKS[workbook_id] = {
        'workbook': workbook,
        'filename': filename,
        'library': 'xlsxwriter',
        'worksheets': {},
        'formats': {}
    }
    
    return {
        'success': True,
        'result': {
            'workbook_id': workbook_id,
            'library': 'xlsxwriter',
            'filename': filename
        }
    }

def _create_workbook_basic(filename: str) -> Dict[str, Any]:
    """Create basic workbook (fallback - CSV-like)"""
    workbook_id = str(uuid.uuid4())
    WORKBOOKS[workbook_id] = {
        'workbook': None,
        'filename': filename,
        'library': 'basic',
        'worksheets': {},
        'formats': {},
        'data': []  # Store data for CSV output
    }
    
    return {
        'success': True,
        'result': {
            'workbook_id': workbook_id,
            'library': 'basic',
            'filename': filename
        }
    }

def add_worksheet(workbook_id: str, name: str = '') -> Dict[str, Any]:
    """
    Add worksheet to workbook
    
    Args:
        workbook_id: Workbook identifier
        name: Worksheet name (optional)
    
    Returns:
        Dictionary with worksheet creation result
    """
    try:
        if workbook_id not in WORKBOOKS:
            return {
                'success': False,
                'error': 'Workbook not found'
            }
        
        wb_data = WORKBOOKS[workbook_id]
        library = wb_data['library']
        
        if library == 'openpyxl':
            return _add_worksheet_openpyxl(workbook_id, name)
        elif library == 'xlsxwriter':
            return _add_worksheet_xlsxwriter(workbook_id, name)
        else:
            return _add_worksheet_basic(workbook_id, name)
            
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to add worksheet: {str(e)}'
        }

def _add_worksheet_openpyxl(workbook_id: str, name: str) -> Dict[str, Any]:
    """Add worksheet using openpyxl"""
    wb_data = WORKBOOKS[workbook_id]
    workbook = wb_data['workbook']
    
    worksheet = workbook.create_sheet(title=name or 'Sheet1')
    
    worksheet_id = str(uuid.uuid4())
    wb_data['worksheets'][worksheet_id] = {
        'worksheet': worksheet,
        'name': name or 'Sheet1'
    }
    
    return {
        'success': True,
        'result': {
            'worksheet_id': worksheet_id,
            'name': name or 'Sheet1'
        }
    }

def _add_worksheet_xlsxwriter(workbook_id: str, name: str) -> Dict[str, Any]:
    """Add worksheet using xlsxwriter"""
    wb_data = WORKBOOKS[workbook_id]
    workbook = wb_data['workbook']
    
    worksheet = workbook.add_worksheet(name or None)
    
    worksheet_id = str(uuid.uuid4())
    wb_data['worksheets'][worksheet_id] = {
        'worksheet': worksheet,
        'name': name or 'Sheet1'
    }
    
    return {
        'success': True,
        'result': {
            'worksheet_id': worksheet_id,
            'name': name or 'Sheet1'
        }
    }

def _add_worksheet_basic(workbook_id: str, name: str) -> Dict[str, Any]:
    """Add worksheet using basic implementation"""
    wb_data = WORKBOOKS[workbook_id]
    
    worksheet_id = str(uuid.uuid4())
    wb_data['worksheets'][worksheet_id] = {
        'worksheet': {},  # Store as dict of (row,col): value
        'name': name or 'Sheet1',
        'max_row': 0,
        'max_col': 0
    }
    
    return {
        'success': True,
        'result': {
            'worksheet_id': worksheet_id,
            'name': name or 'Sheet1'
        }
    }

def add_format(workbook_id: str, properties: Dict[str, Any] = None) -> Dict[str, Any]:
    """
    Add format to workbook
    
    Args:
        workbook_id: Workbook identifier
        properties: Format properties
    
    Returns:
        Dictionary with format creation result
    """
    try:
        if workbook_id not in WORKBOOKS:
            return {
                'success': False,
                'error': 'Workbook not found'
            }
        
        wb_data = WORKBOOKS[workbook_id]
        library = wb_data['library']
        properties = properties or {}
        
        if library == 'openpyxl':
            return _add_format_openpyxl(workbook_id, properties)
        elif library == 'xlsxwriter':
            return _add_format_xlsxwriter(workbook_id, properties)
        else:
            return _add_format_basic(workbook_id, properties)
            
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to add format: {str(e)}'
        }

def _add_format_openpyxl(workbook_id: str, properties: Dict[str, Any]) -> Dict[str, Any]:
    """Add format using openpyxl"""
    wb_data = WORKBOOKS[workbook_id]
    
    format_id = str(uuid.uuid4())
    wb_data['formats'][format_id] = {
        'properties': properties.copy(),
        'library': 'openpyxl'
    }
    
    return {
        'success': True,
        'result': {
            'format_id': format_id,
            'properties': properties
        }
    }

def _add_format_xlsxwriter(workbook_id: str, properties: Dict[str, Any]) -> Dict[str, Any]:
    """Add format using xlsxwriter"""
    wb_data = WORKBOOKS[workbook_id]
    workbook = wb_data['workbook']
    
    # Create xlsxwriter format
    xl_format = workbook.add_format()
    
    # Apply properties
    _apply_xlsxwriter_properties(xl_format, properties)
    
    format_id = str(uuid.uuid4())
    wb_data['formats'][format_id] = {
        'format': xl_format,
        'properties': properties.copy(),
        'library': 'xlsxwriter'
    }
    
    return {
        'success': True,
        'result': {
            'format_id': format_id,
            'properties': properties
        }
    }

def _add_format_basic(workbook_id: str, properties: Dict[str, Any]) -> Dict[str, Any]:
    """Add format using basic implementation"""
    wb_data = WORKBOOKS[workbook_id]
    
    format_id = str(uuid.uuid4())
    wb_data['formats'][format_id] = {
        'properties': properties.copy(),
        'library': 'basic'
    }
    
    return {
        'success': True,
        'result': {
            'format_id': format_id,
            'properties': properties
        }
    }

def write_cell(workbook_id: str, worksheet_id: str, row: int, col: int, 
               data: Any, format_id: str = None) -> Dict[str, Any]:
    """
    Write data to a cell
    
    Args:
        workbook_id: Workbook identifier
        worksheet_id: Worksheet identifier
        row: Row number (0-based)
        col: Column number (0-based)
        data: Data to write
        format_id: Format identifier (optional)
    
    Returns:
        Dictionary with write operation result
    """
    try:
        if workbook_id not in WORKBOOKS:
            return {
                'success': False,
                'error': 'Workbook not found'
            }
        
        wb_data = WORKBOOKS[workbook_id]
        library = wb_data['library']
        
        if worksheet_id not in wb_data['worksheets']:
            return {
                'success': False,
                'error': 'Worksheet not found'
            }
        
        if library == 'openpyxl':
            return _write_cell_openpyxl(workbook_id, worksheet_id, row, col, data, format_id)
        elif library == 'xlsxwriter':
            return _write_cell_xlsxwriter(workbook_id, worksheet_id, row, col, data, format_id)
        else:
            return _write_cell_basic(workbook_id, worksheet_id, row, col, data, format_id)
            
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to write cell: {str(e)}'
        }

def _write_cell_openpyxl(workbook_id: str, worksheet_id: str, row: int, col: int, 
                        data: Any, format_id: str = None) -> Dict[str, Any]:
    """Write cell using openpyxl"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb_data = WORKBOOKS[workbook_id]
    ws_data = wb_data['worksheets'][worksheet_id]
    worksheet = ws_data['worksheet']
    
    # Write data (openpyxl uses 1-based indexing)
    cell = worksheet.cell(row=row+1, column=col+1, value=data)
    
    # Apply format if provided
    if format_id and format_id in wb_data['formats']:
        format_props = wb_data['formats'][format_id]['properties']
        _apply_openpyxl_format(cell, format_props)
    
    return {
        'success': True,
        'result': {
            'row': row,
            'col': col,
            'data': data,
            'formatted': bool(format_id)
        }
    }

def _write_cell_xlsxwriter(workbook_id: str, worksheet_id: str, row: int, col: int, 
                          data: Any, format_id: str = None) -> Dict[str, Any]:
    """Write cell using xlsxwriter"""
    wb_data = WORKBOOKS[workbook_id]
    ws_data = wb_data['worksheets'][worksheet_id]
    worksheet = ws_data['worksheet']
    
    # Get format if provided
    xl_format = None
    if format_id and format_id in wb_data['formats']:
        xl_format = wb_data['formats'][format_id]['format']
    
    # Write data
    worksheet.write(row, col, data, xl_format)
    
    return {
        'success': True,
        'result': {
            'row': row,
            'col': col,
            'data': data,
            'formatted': bool(format_id)
        }
    }

def _write_cell_basic(workbook_id: str, worksheet_id: str, row: int, col: int, 
                     data: Any, format_id: str = None) -> Dict[str, Any]:
    """Write cell using basic implementation"""
    wb_data = WORKBOOKS[workbook_id]
    ws_data = wb_data['worksheets'][worksheet_id]
    worksheet = ws_data['worksheet']
    
    # Store data
    worksheet[(row, col)] = {
        'data': data,
        'format_id': format_id
    }
    
    # Update max dimensions
    ws_data['max_row'] = max(ws_data['max_row'], row)
    ws_data['max_col'] = max(ws_data['max_col'], col)
    
    return {
        'success': True,
        'result': {
            'row': row,
            'col': col,
            'data': data,
            'formatted': bool(format_id)
        }
    }

def update_format(workbook_id: str, format_id: str, property: str, value: Any) -> Dict[str, Any]:
    """
    Update format property
    
    Args:
        workbook_id: Workbook identifier
        format_id: Format identifier
        property: Property name
        value: Property value
    
    Returns:
        Dictionary with update result
    """
    try:
        if workbook_id not in WORKBOOKS:
            return {
                'success': False,
                'error': 'Workbook not found'
            }
        
        wb_data = WORKBOOKS[workbook_id]
        
        if format_id not in wb_data['formats']:
            return {
                'success': False,
                'error': 'Format not found'
            }
        
        # Update property
        wb_data['formats'][format_id]['properties'][property] = value
        
        # If xlsxwriter, update the actual format object
        if wb_data['library'] == 'xlsxwriter':
            xl_format = wb_data['formats'][format_id]['format']
            _apply_xlsxwriter_property(xl_format, property, value)
        
        return {
            'success': True,
            'result': {
                'format_id': format_id,
                'property': property,
                'value': value
            }
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to update format: {str(e)}'
        }

def close_workbook(workbook_id: str) -> Dict[str, Any]:
    """
    Close and save workbook
    
    Args:
        workbook_id: Workbook identifier
    
    Returns:
        Dictionary with close operation result
    """
    try:
        if workbook_id not in WORKBOOKS:
            return {
                'success': False,
                'error': 'Workbook not found'
            }
        
        wb_data = WORKBOOKS[workbook_id]
        library = wb_data['library']
        
        if library == 'openpyxl':
            return _close_workbook_openpyxl(workbook_id)
        elif library == 'xlsxwriter':
            return _close_workbook_xlsxwriter(workbook_id)
        else:
            return _close_workbook_basic(workbook_id)
            
    except Exception as e:
        return {
            'success': False,
            'error': f'Failed to close workbook: {str(e)}'
        }
    finally:
        # Clean up memory
        if workbook_id in WORKBOOKS:
            del WORKBOOKS[workbook_id]

def _close_workbook_openpyxl(workbook_id: str) -> Dict[str, Any]:
    """Close workbook using openpyxl"""
    wb_data = WORKBOOKS[workbook_id]
    workbook = wb_data['workbook']
    filename = wb_data['filename']
    
    workbook.save(filename)
    
    return {
        'success': True,
        'result': {
            'filename': filename,
            'library': 'openpyxl'
        }
    }

def _close_workbook_xlsxwriter(workbook_id: str) -> Dict[str, Any]:
    """Close workbook using xlsxwriter"""
    wb_data = WORKBOOKS[workbook_id]
    workbook = wb_data['workbook']
    filename = wb_data['filename']
    
    workbook.close()
    
    return {
        'success': True,
        'result': {
            'filename': filename,
            'library': 'xlsxwriter'
        }
    }

def _close_workbook_basic(workbook_id: str) -> Dict[str, Any]:
    """Close workbook using basic implementation (save as CSV)"""
    import csv
    
    wb_data = WORKBOOKS[workbook_id]
    filename = wb_data['filename']
    
    # Convert .xlsx to .csv for basic implementation
    if filename.endswith('.xlsx'):
        filename = filename[:-5] + '.csv'
    
    # Save first worksheet as CSV
    if wb_data['worksheets']:
        first_ws = list(wb_data['worksheets'].values())[0]
        worksheet = first_ws['worksheet']
        max_row = first_ws['max_row']
        max_col = first_ws['max_col']
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            for row in range(max_row + 1):
                row_data = []
                for col in range(max_col + 1):
                    cell_data = worksheet.get((row, col), {})
                    row_data.append(str(cell_data.get('data', '')))
                writer.writerow(row_data)
    
    return {
        'success': True,
        'result': {
            'filename': filename,
            'library': 'basic'
        }
    }

# Helper functions for format application

def _apply_openpyxl_format(cell, properties: Dict[str, Any]) -> None:
    """Apply format properties to openpyxl cell"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Font properties
    font_kwargs = {}
    if 'bold' in properties:
        font_kwargs['bold'] = properties['bold']
    if 'font_color' in properties:
        font_kwargs['color'] = _convert_color(properties['font_color'])
    
    if font_kwargs:
        cell.font = Font(**font_kwargs)
    
    # Fill (background color)
    if 'bg_color' in properties:
        fill_color = _convert_color(properties['bg_color'])
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    # Alignment
    if 'align' in properties:
        alignment = properties['align']
        if alignment == 'center':
            cell.alignment = Alignment(horizontal='center')
        elif alignment == 'left':
            cell.alignment = Alignment(horizontal='left')
        elif alignment == 'right':
            cell.alignment = Alignment(horizontal='right')

def _apply_xlsxwriter_properties(xl_format, properties: Dict[str, Any]) -> None:
    """Apply all properties to xlsxwriter format"""
    for prop, value in properties.items():
        _apply_xlsxwriter_property(xl_format, prop, value)

def _apply_xlsxwriter_property(xl_format, property: str, value: Any) -> None:
    """Apply single property to xlsxwriter format"""
    if property == 'bold':
        xl_format.set_bold(value)
    elif property == 'font_color':
        xl_format.set_font_color(_convert_color(value))
    elif property == 'bg_color':
        xl_format.set_bg_color(_convert_color(value))
    elif property == 'align':
        xl_format.set_align(value)
    elif property == 'border':
        xl_format.set_border(value)
    elif property == 'num_format':
        xl_format.set_num_format(value)

def _convert_color(color: str) -> str:
    """Convert color name to hex or keep as-is"""
    color_map = {
        'black': '000000',
        'white': 'FFFFFF',
        'gray': 'C0C0C0',
        'grey': 'C0C0C0',
        'red': 'FF0000',
        'green': '00FF00',
        'blue': '0000FF',
        'yellow': 'FFFF00',
    }
    
    return color_map.get(color.lower(), color)

# Test functions
def ping():
    """Basic connectivity test"""
    return {
        'message': 'Excel backend is ready',
        'supported_methods': [
            'create_workbook', 'add_worksheet', 'add_format',
            'write_cell', 'update_format', 'close_workbook'
        ],
        'version': '1.0.0',
        'libraries': _get_available_libraries(),
        'active_workbooks': len(WORKBOOKS)
    }

def _get_available_libraries():
    """Check which Excel libraries are available"""
    libraries = []
    
    try:
        import openpyxl
        libraries.append(f'openpyxl {openpyxl.__version__}')
    except ImportError:
        pass
    
    try:
        import xlsxwriter
        libraries.append(f'xlsxwriter {xlsxwriter.__version__}')
    except ImportError:
        pass
    
    libraries.append('basic (CSV fallback)')
    
    return libraries

def test_excel_creation():
    """Test Excel creation with your usage pattern"""
    try:
        # Test workbook creation
        wb_result = create_workbook('/tmp/test_excel.xlsx')
        if not wb_result['success']:
            return wb_result
        
        workbook_id = wb_result['result']['workbook_id']
        
        # Test worksheet creation
        ws_result = add_worksheet(workbook_id, 'TestSheet')
        if not ws_result['success']:
            return ws_result
        
        worksheet_id = ws_result['result']['worksheet_id']
        
        # Test format creation (your header format pattern)
        fmt_result = add_format(workbook_id, {
            'bold': True,
            'font_color': 'black',
            'bg_color': 'gray',
            'align': 'center'
        })
        if not fmt_result['success']:
            return fmt_result
        
        format_id = fmt_result['result']['format_id']
        
        # Test cell writing (your pattern)
        headers = ['Name', 'Age', 'City']
        for x, header in enumerate(headers):
            write_result = write_cell(workbook_id, worksheet_id, 0, x, header, format_id)
            if not write_result['success']:
                return write_result
        
        # Test data writing
        test_data = [
            {'Name': 'John', 'Age': 30, 'City': 'New York'},
            {'Name': 'Jane', 'Age': 25, 'City': 'Chicago'}
        ]
        
        for y, row in enumerate(test_data, 1):
            for x, header in enumerate(headers):
                write_result = write_cell(workbook_id, worksheet_id, y, x, row[header])
                if not write_result['success']:
                    return write_result
        
        # Test close
        close_result = close_workbook(workbook_id)
        if not close_result['success']:
            return close_result
        
        return {
            'success': True,
            'test': 'excel_creation',
            'result': {
                'workbook_created': True,
                'worksheet_created': True,
                'format_created': True,
                'headers_written': len(headers),
                'data_rows_written': len(test_data),
                'file_saved': close_result['result']['filename'],
                'library_used': wb_result['result']['library']
            }
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Excel creation test failed: {str(e)}'
        }

def simulate_exporttoexcel(data: list, filename: str) -> Dict[str, Any]:
    """
    Simulate your exact exportToExcel subroutine pattern
    
    Args:
        data: List of dictionaries (your $self->data() pattern)
        filename: Output filename
    
    Returns:
        Dictionary with operation result
    """
    try:
        if not data or not isinstance(data[0], dict):
            return {
                'success': False,
                'error': 'Data must be list of dictionaries'
            }
        
        # Create workbook
        wb_result = create_workbook(filename)
        if not wb_result['success']:
            return wb_result
        
        workbook_id = wb_result['result']['workbook_id']
        
        # Add worksheet
        ws_result = add_worksheet(workbook_id)
        if not ws_result['success']:
            return ws_result
        
        worksheet_id = ws_result['result']['worksheet_id']
        
        # Create header format (your exact pattern)
        fmt_result = add_format(workbook_id, {
            'bold': True,
            'font_color': 'black', 
            'bg_color': 'gray',
            'align': 'center'
        })
        if not fmt_result['success']:
            return fmt_result
        
        format_id = fmt_result['result']['format_id']
        
        # Write headers (your pattern: sorted keys)
        keys = sorted(data[0].keys())
        x = y = 0
        
        for header in keys:
            write_result = write_cell(workbook_id, worksheet_id, y, x, header, format_id)
            if not write_result['success']:
                return write_result
            x += 1
        
        # Write data rows (your pattern)
        for row in data:
            x = 0
            y += 1
            for key in keys:
                write_result = write_cell(workbook_id, worksheet_id, y, x, row.get(key, ''))
                if not write_result['success']:
                    return write_result
                x += 1
        
        # Close workbook
        close_result = close_workbook(workbook_id)
        
        return {
            'success': True,
            'result': {
                'filename': filename,
                'headers': keys,
                'rows_written': len(data),
                'library': wb_result['result']['library']
            }
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'exportToExcel simulation failed: {str(e)}'
        }